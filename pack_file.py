import os
import csv

from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook


def write_and_read_zip():
    if os.path.exists('resources/arhive.zip'):
        os.remove('resources/arhive.zip')

    with ZipFile('resources/arhive.zip', mode='w') as myzip:
        print('start zip file')
        myzip.write('./sample.pdf')
        myzip.write('./file_example_XLSX_50.xlsx')
        myzip.write('./username.csv')
        print('complete zip file')


    with ZipFile('resources/arhive.zip', mode='r') as zip_file:
        arhived_file = zip_file.namelist()
        print(f'В архиве находятся {arhived_file}')
        with zip_file.open('sample.pdf') as mypdf:
            reader = PdfReader(mypdf)
            pdf_text = reader.pages[0].extractText()
            assert 'A Simple PDF File' in pdf_text, 'В PDF файле нет такой строки'

        with zip_file.open('file_example_XLSX_50.xlsx') as myxlsx:
            workbook = load_workbook(myxlsx)
            sheet = workbook.active
            print(sheet.cell(row=3, column=2).value)
            assert 'Mara' == sheet.cell(row=3, column=2).value, 'В заданной ячейке не содержится "Mara"'

        with open('username.csv') as mycsv:
            csvreader = csv.DictReader(mycsv)
            for row in csvreader:
                print(row)
            assert 'Username' in csvreader.fieldnames[0], 'Поля "Username" нет в таблице'



if __name__ == '__main__':
    write_and_read_zip()